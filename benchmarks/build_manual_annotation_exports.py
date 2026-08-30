from __future__ import annotations

import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
ANNOTATED = ROOT / "benchmarks/datasets/teacher_math_manual_annotated_20260827.json"
SNAPSHOT = ROOT / "output/benchmarks/teacher-file-oracle-20260826-rerun/source_snapshot.json"
OUT = ROOT / "benchmarks/datasets"

# These are manually authored semantic prompts. The script never derives a query
# from source text; it only binds a reviewed path/order to the current snapshot.
REVIEWED_CORRECTIONS = {
    "AG-007": {
        "expected_block_order": 6,
        "target_section": "椭圆的第二定义",
    },
    "AG-012": {
        "expected_block_order": 7,
        "target_section": "3.1.1 过焦点坐标相乘为定值",
    },
    "AG-013": {
        "expected_block_order": 5,
        "target_section": "2.抛物线的通径",
    },
    "PR-005": {
        "expected_block_order": 4,
        "target_section": "一、有放回",
    },
    "PR-011": {
        "expected_block_order": 5,
        "target_section": "(2) 递推关系 + 通项",
        "query": "累计得分由最后一次加1分或加2分决定时，概率递推关系怎样建立并求出通项？",
        "query_intent": "递推关系与通项",
        "manual_summary": "应按最后一步加1分或加2分分类，建立B_n=(1/3)B_{n-1}+(2/3)B_{n-2}并说明初值与通项。",
        "acceptable_evidence": ["最后一步分类", "递推式", "初值与通项"],
    },
    "SG-001": {
        "expected_block_order": 2,
        "target_section": "① 线面平行判定",
    },
    "SG-002": {
        "expected_block_order": 6,
        "target_section": "③ 面面平行判定",
    },
    "SG-003": {
        "expected_block_order": 12,
        "target_section": "① 线面垂直判定",
    },
    "SG-009": {
        "expected_block_order": 3,
        "target_section": "球的垂径定理",
    },
}

ADDITIONS = [
    {"case_id":"AG-015","topic":"解析几何/点与椭圆","query_intent":"点在曲线上判断","query":"给定一个点的坐标，怎样快速判断它在椭圆上、椭圆内还是椭圆外？","expected_source_path":"解析几何/点与圆锥曲线、直线位置关系.md","expected_block_order":3,"target_section":"三、点与椭圆","manual_summary":"应代入椭圆左端表达式，与1比较判断位置。","acceptable_evidence":["代入椭圆方程","等于1在椭圆上","与1比较"]},
    {"case_id":"AG-016","topic":"解析几何/点与抛物线","query_intent":"开口方向位置判断","query":"点和抛物线的位置关系判断，开口向右与开口向上的代入变量为什么不同？","expected_source_path":"解析几何/点与圆锥曲线、直线位置关系.md","expected_block_order":5,"target_section":"1. 上下开口（代x比y）","manual_summary":"应说明上下开口代x比较y，左右开口代y比较x，并解释内外侧方向。","acceptable_evidence":["上下开口代x","左右开口代y","内外侧判断"]},
    {"case_id":"AG-017","topic":"解析几何/角度转化","query_intent":"斜率与向量选择","query":"解析几何中的角度条件，什么时候用斜率，什么时候用向量点积更稳妥？","expected_source_path":"解析几何/怎么翻译/直线角度要翻译成向量_斜率.md","expected_block_order":8,"target_section":"1. 优先使用向量的场景","manual_summary":"应区分斜率存在与否、顶点位置和不规则角，说明向量的兜底作用。","acceptable_evidence":["斜率不存在","角顶点不在坐标轴","向量兜底"]},
    {"case_id":"AG-018","topic":"解析几何/角度转化","query_intent":"垂直条件翻译","query":"两条直线斜率都存在时，垂直条件怎样翻译成代数式？竖直线要怎么补充？","expected_source_path":"解析几何/怎么翻译/直线角度要翻译成向量_斜率.md","expected_block_order":5,"target_section":"模块1：垂直/直角的斜率转化","manual_summary":"应给出 k1k2=-1，并说明斜率不存在时需单独讨论水平线与竖直线。","acceptable_evidence":["k1k2=-1","斜率均存在","水平竖直特殊情形"]},
    {"case_id":"AG-019","topic":"解析几何/轨迹","query_intent":"参数消元轨迹","query":"由椭圆上的动点构造中点时，怎样用参数式消去参数得到中点轨迹？","expected_source_path":"解析几何/求点轨迹问题.md","expected_block_order":1,"target_section":"构造点轨迹的参数消元","manual_summary":"应先写构造点坐标，再利用参数关系消去t，得到轨迹方程。","acceptable_evidence":["构造点坐标","参数消元","轨迹方程"]},
    {"case_id":"AG-020","topic":"解析几何/定值定点","query_intent":"比例式定值","query":"解析几何里一个分式要恒定，为什么常常要把分子分母化成固定比例？","expected_source_path":"解析几何/定值定点/定值是怎么来的？.md","expected_block_order":0,"target_section":"定值是怎么来的","manual_summary":"应说明通分、因式分解或消元后比较分子分母的比例。","acceptable_evidence":["分子分母固定比例","通分化简","定值"]},
    {"case_id":"AG-021","topic":"解析几何/焦点三角形","query_intent":"内心坐标","query":"椭圆焦点三角形的内心横坐标为什么可以用焦半径和三角形边长表示？","expected_source_path":"解析几何/焦点三角形内心坐标公式.md","expected_block_order":2,"target_section":"一、椭圆焦点三角形内心坐标推导","manual_summary":"应利用内心按边长加权的坐标公式，并代入两焦半径。","acceptable_evidence":["内心坐标加权","焦半径","椭圆焦点三角形"]},
    {"case_id":"SG-013","topic":"立体几何/三余弦定理","query_intent":"最小角定理","query":"为什么斜线与平面内直线的夹角不小于线面角？","expected_source_path":"立体几何/三垂弦定理三余弦定理_三正弦定理 知识点.md","expected_block_order":8,"target_section":"线面角是最小角","manual_summary":"应由 cosθ2≤1 和余弦单调性推出 φ≥θ。","acceptable_evidence":["线面角最小","cosθ2≤1","余弦单调"]},
    {"case_id":"SG-014","topic":"立体几何/三垂线","query_intent":"逆定理","query":"已知斜线垂直平面内一条直线，怎样推出它的射影也垂直这条直线？","expected_source_path":"立体几何/立体几何平行垂直基本知识.md","expected_block_order":24,"target_section":"二、三垂线逆定理","manual_summary":"应识别三垂线逆定理，区分斜线、射影和目标平面内直线。","acceptable_evidence":["三垂线逆定理","斜线垂直","射影垂直"]},
    {"case_id":"SG-015","topic":"立体几何/建系","query_intent":"墙角模型","query":"三个面两两垂直的墙角模型，空间直角坐标系的原点和三条轴怎样选？","expected_source_path":"立体几何/立体几何常见模型/墙角模型.md","expected_block_order":1,"target_section":"墙角模型建系","manual_summary":"应说明取墙角顶点为原点，沿三条互相垂直棱设轴。","acceptable_evidence":["墙角三面垂直","原点","三条坐标轴"]},
    {"case_id":"SG-016","topic":"立体几何/平面图","query_intent":"三维二维化","query":"立体几何图形关系太复杂时，为什么要抽取一个平面截面来分析？","expected_source_path":"立体几何/立体几何小题/TODO 三维问题二维化 立体几何画出平面图.md","expected_block_order":0,"target_section":"三维问题二维化","manual_summary":"应说明选取包含关键线段和垂直关系的平面，把空间关系转成平面几何。","acceptable_evidence":["选取平面截面","保留关键线段","空间转平面"]},
    {"case_id":"SG-017","topic":"立体几何/外接球","query_intent":"正棱锥模型","query":"正棱锥外接球计算中，为什么球心、底面外心和顶点可以放进同一个截面？","expected_source_path":"立体几何/外接球和内切球/外接球常见模型 长方体 棱柱 棱锥.md","expected_block_order":2,"target_section":"正棱锥外接球半径推导","manual_summary":"应说明正棱锥对称性使球心落在高线上，截面得到直角三角形。","acceptable_evidence":["对称轴","底面外心","截面直角三角形"]},
    {"case_id":"SG-018","topic":"立体几何/体积","query_intent":"体积公式","query":"棱柱和棱锥的体积公式分别依赖底面积和什么高度关系？","expected_source_path":"立体几何/常见立体几何面积表面积/棱台和棱柱的体积.md","expected_block_order":1,"target_section":"棱柱和棱台体积","manual_summary":"应区分棱柱底面积乘高与棱锥三分之一底面积乘高。","acceptable_evidence":["棱柱底面积乘高","棱锥三分之一","垂直高"]},
    {"case_id":"SQ-009","topic":"数列/等比数列","query_intent":"等比求和条件","query":"等比数列前n项和公式为什么必须区分公比q等于1和不等于1？","expected_source_path":"数列/数列求和/基本数列求和公式.md","expected_block_order":4,"target_section":"等比数列求和公式","manual_summary":"应给出q≠1的分式公式，并说明q=1时直接为na1。","acceptable_evidence":["q≠1","等比求和","q=1特殊情形"]},
    {"case_id":"SQ-010","topic":"数列/分组求和","query_intent":"通项拆分","query":"通项由等比项、等差项和常数相加组成时，前n项和怎样分组计算？","expected_source_path":"数列/数列求和/基本数列求和公式.md","expected_block_order":9,"target_section":"等差、等比与常数分组求和","manual_summary":"应把通项拆成三部分，分别使用等比、等差和常数求和。","acceptable_evidence":["通项拆分","分组求和","三类基本公式"]},
    {"case_id":"SQ-011","topic":"数列/递推","query_intent":"前n项和递推","query":"递推式同时含有 a(n+1) 和 Sn 时，怎样先用 Sn-S(n-1) 把它改写成通项递推？","expected_source_path":"数列/数列求和/基本数列求和公式.md","expected_block_order":12,"target_section":"含Sn递推求通项","manual_summary":"应先处理a1和Sn关系，再消去Sn得到一阶递推或等比结构。","acceptable_evidence":["Sn-S(n-1)","先求a1","递推化简"]},
    {"case_id":"SQ-012","topic":"数列/等差数列","query_intent":"等差不等式","query":"等差数列前n项和的不等式条件，怎样转化成公差正负来判断单调性？","expected_source_path":"数列/等差数列经典结论练习题.md","expected_block_order":2,"target_section":"解题推导","manual_summary":"应通过相邻前n项和之差得到相邻项比较，再判断公差。","acceptable_evidence":["相邻前n项和差","相邻项比较","公差正负"]},
    {"case_id":"SQ-013","topic":"数列/等差数列","query_intent":"特殊值法","query":"等差数列只有比例条件而没有具体首项时，为什么可以令首项为1来求某些比值？","expected_source_path":"数列/等差数列经典结论练习题.md","expected_block_order":11,"target_section":"方法一：特殊值法","manual_summary":"应说明比值对整体缩放不变，特殊赋值可避免求出首项具体值。","acceptable_evidence":["特殊值法","整体缩放","比值不变"]},
    {"case_id":"LC-006","topic":"直线与圆/直线角度","query_intent":"两直线夹角","query":"两条直线斜率已知时，夹角正切公式为什么要取绝对值？","expected_source_path":"直线与圆/直线/TODO 补充图片 到角公式.md","expected_block_order":2,"target_section":"两条直线的夹角公式","manual_summary":"应给出斜率夹角公式，并说明几何夹角取非负值。","acceptable_evidence":["斜率夹角公式","绝对值","几何夹角非负"]},
    {"case_id":"LC-007","topic":"直线与圆/直线角度","query_intent":"平行垂直判定","query":"斜率存在时，两直线平行和垂直分别对应什么代数条件？","expected_source_path":"直线与圆/直线/TODO 补充图片 到角公式.md","expected_block_order":10,"target_section":"总结","manual_summary":"应同时返回k1=k2与k1k2=-1，并提醒竖直线需单独判断。","acceptable_evidence":["k1=k2","k1k2=-1","竖直线特殊"]},
    {"case_id":"LC-008","topic":"直线与圆/位置关系","query_intent":"点与圆位置","query":"给出圆心和半径后，如何只比较距离平方判断点在圆内、圆上还是圆外？","expected_source_path":"解析几何/点与圆锥曲线、直线位置关系.md","expected_block_order":2,"target_section":"二、点与圆（圆心距法，标准解法）","manual_summary":"应比较点到圆心距离平方与r²，避免不必要开根号。","acceptable_evidence":["圆心距平方","与r²比较","圆内圆上圆外"]},
    {"case_id":"LC-009","topic":"直线与圆/阿波罗尼斯圆","query_intent":"轨迹特殊情形","query":"两定点距离比等于1时，阿波罗尼斯圆为什么退化成垂直平分线？","expected_source_path":"直线与圆/阿波罗尼斯圆 （阿氏圆）.md","expected_block_order":0,"target_section":"阿波罗尼斯圆","manual_summary":"应指出到两定点距离相等的轨迹是垂直平分线，这是比值为1的特殊情形。","acceptable_evidence":["距离相等","垂直平分线","比值1"]},
    {"case_id":"LC-010","topic":"直线与圆/角度向量","query_intent":"向量点积求角","query":"角的顶点不在坐标轴上时，怎样用两条方向向量的点积求夹角？","expected_source_path":"解析几何/怎么翻译/直线角度要翻译成向量_斜率.md","expected_block_order":9,"target_section":"2. 核心转化公式（万能无限制）","manual_summary":"应给出点积除以模长乘积的余弦公式和向量取向。","acceptable_evidence":["方向向量","点积","模长乘积"]},
    {"case_id":"PR-013","topic":"概率统计/均值","query_intent":"均值与总和","query":"知道样本容量和平均数后，怎样把均值条件改写成数据总和条件？","expected_source_path":"概率统计/数据统计/TODO 均值的表示和变换.md","expected_block_order":1,"target_section":"题目里均值的提取通用步骤","manual_summary":"应使用Σxi=n xbar，并能处理拆分后的求和式。","acceptable_evidence":["Σxi=n xbar","样本容量","拆分求和"]},
    {"case_id":"PR-014","topic":"概率统计/随机变量","query_intent":"随机变量线性变换","query":"已知随机变量X的期望和方差，Y=aX+b的期望方差怎样直接得到？","expected_source_path":"概率统计/数据统计/期望和方差/期望和方差的性质.md","expected_block_order":1,"target_section":"期望和方差的性质","manual_summary":"应给出E(aX+b)=aE(X)+b和D(aX+b)=a²D(X)。","acceptable_evidence":["期望线性变换","方差平方系数","常数平移"]},
    {"case_id":"PR-015","topic":"概率统计/古典概型","query_intent":"等可能建模","query":"古典概型成立需要哪些基本条件，为什么有利结果数除以总结果数？","expected_source_path":"概率统计/AI 古典概型.md","expected_block_order":0,"target_section":"古典概型","manual_summary":"应说明有限样本空间、基本事件等可能和计数比值。","acceptable_evidence":["有限样本空间","等可能基本事件","有利数除总数"]},
    {"case_id":"PR-016","topic":"概率统计/抽签","query_intent":"无放回条件概率","query":"依次不放回抽签时，前面的人没有抽中特殊签，后面抽中的概率怎样更新？","expected_source_path":"概率统计/概率模型/抽签的概率是相同的.md","expected_block_order":2,"target_section":"解答","manual_summary":"应根据条件发生后的剩余签数和特殊签数重新计算，而不是把每次概率当独立。","acceptable_evidence":["不放回","条件更新","剩余签数"]},
    {"case_id":"PR-017","topic":"概率统计/二项分布","query_intent":"独立重复试验","query":"二项分布模型需要哪些独立重复条件，随机变量表示的到底是什么次数？","expected_source_path":"概率统计/数据统计/分布问题/TODO 二项分布.md","expected_block_order":0,"target_section":"二项分布","manual_summary":"应检查试验次数固定、每次成功概率相同且相互独立，并明确X为成功次数。","acceptable_evidence":["固定次数","相同成功概率","独立","成功次数"]},
    {"case_id":"PR-018","topic":"概率统计/期望方差","query_intent":"方差推导","query":"从方差定义展开完全平方后，哪一步能得到E(X²)-[E(X)]²？","expected_source_path":"概率统计/数据统计/期望和方差/期望和方差的性质.md","expected_block_order":4,"target_section":"步骤2：两边取期望","manual_summary":"应强调E(X)是常数并使用期望线性性质合并各项。","acceptable_evidence":["E(X)为常数","期望线性","E(X²)-[E(X)]²"]},
    {"case_id":"TR-011","topic":"解三角形/射影定理","query_intent":"直角三角形射影","query":"直角三角形斜边上的高把两条直角边分成射影后，平方关系分别是什么？","expected_source_path":"解三角形/三角形常用定理/TODO 射影定理.md","expected_block_order":1,"target_section":"一、定理内容","manual_summary":"应给出两直角边平方等于对应射影乘斜边，高平方等于两射影乘积。","acceptable_evidence":["直角三角形","直角边平方","射影乘斜边"]},
    {"case_id":"TR-012","topic":"解三角形/正弦定理","query_intent":"外接圆联系","query":"正弦定理中的2R为什么是三角形外接圆直径？","expected_source_path":"解三角形/三角形常用定理/正弦定理.md","expected_block_order":0,"target_section":"正弦定理","manual_summary":"应联系边与对角正弦比以及外接圆半径R。","acceptable_evidence":["a/sinA","外接圆半径R","直径2R"]},
    {"case_id":"TR-013","topic":"解三角形/练习题","query_intent":"三角恒等变形","query":"锐角三角形中出现cos2A+cos2B时，为什么可以优先改写成sin²项？","expected_source_path":"解三角形/解三角形练习题/三角函数类的练习题.md","expected_block_order":2,"target_section":"思路","manual_summary":"应说明利用cos2α=1-2sin²α，结合A+B+C=π整理。","acceptable_evidence":["cos2α公式","sin²改写","三角形内角和"]},
    {"case_id":"TR-014","topic":"解三角形/条件转化","query_intent":"射影关系","query":"三角形边长条件中出现 c=b cosA+a cosB，几何上对应什么投影关系？","expected_source_path":"解三角形/TODO 补充 解三角形常见转化条件.md","expected_block_order":1,"target_section":"射影定理","manual_summary":"应解释两条边在第三边方向上的投影相加得到另一边。","acceptable_evidence":["投影","b cosA+a cosB","边长转化"]},
    {"case_id":"TR-015","topic":"解三角形/多选题","query_intent":"正弦定理判断","query":"在三角形多选题中，a cosB=b cosA能推出什么对称的边角关系？","expected_source_path":"解三角形/解三角形练习题/解三角形多选题.md","expected_block_order":1,"target_section":"2026辽宁名校三月","manual_summary":"应结合正弦定理或积化关系判断等腰条件，不能直接凭形式猜结论。","acceptable_evidence":["a cosB=b cosA","正弦定理转化","等腰判断"]},
    {"case_id":"TR-016","topic":"解三角形/三角函数练习","query_intent":"锐角约束","query":"锐角三角形的面积范围题，哪些角度和正弦取值约束必须同时使用？","expected_source_path":"解三角形/解三角形练习题/三角函数类的练习题.md","expected_block_order":1,"target_section":"题目1","manual_summary":"应结合锐角范围、内角和以及正余弦恒等式建立面积范围。","acceptable_evidence":["锐角范围","内角和","面积与sin"]},
    {"case_id":"TG-009","topic":"三角函数/倒数函数","query_intent":"sec csc cot定义","query":"sec、csc、cot分别是哪个三角函数的倒数，定义域要排除什么值？","expected_source_path":"三角函数/sec，csc，cot都是什么.md","expected_block_order":0,"target_section":"核心速记","manual_summary":"应给出sec=1/cos、csc=1/sin、cot=cos/sin并说明分母不能为0。","acceptable_evidence":["sec=1/cos","csc=1/sin","cot=cos/sin","定义域"]},
    {"case_id":"TG-010","topic":"三角函数/特殊角","query_intent":"常见值","query":"3-4-5直角三角形中，锐角的sin、cos、tan怎样由边长得到？","expected_source_path":"三角函数/常见的三角函数.md","expected_block_order":2,"target_section":"一、最基础：3-4-5三角形","manual_summary":"应按对边、邻边、斜边给出三角函数值，并区分互余角。","acceptable_evidence":["3-4-5","对边邻边斜边","sin cos tan"]},
    {"case_id":"TG-011","topic":"三角函数/和差化积","query_intent":"公式结构","query":"sinα+sinβ化成积时，积中的半和半差分别对应什么角？","expected_source_path":"三角函数/积化和差 和差化积.md","expected_block_order":1,"target_section":"①和差公式","manual_summary":"应说明和差化积中的半和与半差结构及符号。","acceptable_evidence":["sinα+sinβ","半和","半差"]},
    {"case_id":"TG-012","topic":"三角函数/同值关系","query_intent":"余弦同值解","query":"cos x1=cos x2的通解为什么有相差周期和相反对称两种形式？","expected_source_path":"三角函数/！！！三角函数两个值相等如何表示.md","expected_block_order":5,"target_section":"1. 基本形式：cos x1=cos x2","manual_summary":"应给出x2-x1=2kπ或x1+x2=2kπ，并解释余弦偶性。","acceptable_evidence":["cos同值","周期2kπ","偶性对称"]},
    {"case_id":"TG-013","topic":"三角函数/难题","query_intent":"周期参数","query":"函数sin(ωx+φ)的图像题中，等距点和周期信息怎样用来确定ω？","expected_source_path":"三角函数/三角函数难题/三角函数难题.md","expected_block_order":2,"target_section":"难题 2026 武汉三调","manual_summary":"应结合相位差、同一直线几何条件和周期/对称性求ω。","acceptable_evidence":["sin(ωx+φ)","等距点","周期参数ω"]},
    {"case_id":"CP-007","topic":"排列组合/插空法","query_intent":"不相邻计数","query":"若特殊元素不能相邻，先排普通元素后插入时，空位数为什么是n+1？","expected_source_path":"排列组合知识点/AI 和问题 捆绑法和插空法.md","expected_block_order":4,"target_section":"步骤2：计算可插入的空位数","manual_summary":"应说明n个普通元素排成一列产生前、中、后三类共n+1个空位。","acceptable_evidence":["n个普通元素","前中后空位","n+1"]},
    {"case_id":"CP-008","topic":"排列组合/插空法","query_intent":"相同与不同元素","query":"从空位中插入k个特殊元素时，特殊元素不同用排列、相同用组合，为什么？","expected_source_path":"排列组合知识点/AI 和问题 捆绑法和插空法.md","expected_block_order":5,"target_section":"步骤3：从空位里选k个插入特殊元素","manual_summary":"应区分空位选择和特殊元素内部顺序，给出A与C两种情况。","acceptable_evidence":["不同特殊元素排列","相同特殊元素组合","一个空位一个元素"]},
    {"case_id":"CP-009","topic":"排列组合/涂色","query_intent":"容斥去重","query":"涂色分类中同一种情况被两个分支重复计算时，怎样用容斥减掉重复部分？","expected_source_path":"排列组合知识点/七、涂色问题（长时间不考）.md","expected_block_order":7,"target_section":"2013年涂色问题","manual_summary":"应说明先分支计数，再减去交集重复，并保持相邻区域颜色限制。","acceptable_evidence":["分类计数","重复交集","容斥减重"]},
    {"case_id":"CP-010","topic":"排列组合/选座","query_intent":"组合选位","query":"固定座位或方格中只需要选择位置而不需要给位置排序时，为什么用组合数？","expected_source_path":"排列组合知识点/基本的选座问题.md","expected_block_order":0,"target_section":"基本的选座问题","manual_summary":"应区分选择座位与安排对象顺序，说明只选位置时顺序不产生新方案。","acceptable_evidence":["固定位置","只选择位置","组合数"]},
    {"case_id":"CP-011","topic":"排列组合/容斥","query_intent":"补集与容斥边界","query":"排列组合中先数总方案再扣掉违规方案时，怎样检查补集是否完整且没有重复？","expected_source_path":"排列组合知识点/正难则反（补集法）.md","expected_block_order":0,"target_section":"正难则反（补集法）","manual_summary":"应说明补集覆盖全部违规情形，若违规情形交叠则需要容斥校正。","acceptable_evidence":["总方案","补集完整","重复情形容斥"]}
]


def text(value):
    return "" if value is None else str(value).strip()


def normalized(value):
    return " ".join(text(value).replace("[Markdown image block; no extractable text]", "").split())


def split_fingerprint(blocks):
    import hashlib
    parts = []
    for block in blocks:
        body = normalized(block.get("normalizedText") or block.get("rawText"))
        parts.append("|".join((text(block.get("blockType")), text(block.get("blockRole")),
                               text(block.get("pageNo")), text(block.get("blockOrder")),
                               str(len(body)), hashlib.sha256(body.encode("utf-8")).hexdigest())))
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


def bind_case(case, by_path, blocks_by_doc, index):
    path = case["expected_source_path"]
    resource = by_path.get(path)
    if resource is None:
        raise RuntimeError(f"missing real source path for {case['case_id']}: {path}")
    file_id = text(resource.get("documentId"))
    order = int(case["expected_block_order"])
    blocks = blocks_by_doc.get(file_id, [])
    candidates = [block for block in blocks if int(block.get("blockOrder") or 0) == order]
    if not candidates:
        raise RuntimeError(f"missing block order for {case['case_id']}: {path}#{order}")
    block = candidates[0]
    fingerprint = text(resource.get("splitFingerprint")) or split_fingerprint(blocks)
    expected = dict(case)
    expected.update({
        "case_type": "positive",
        "expected_document_id": file_id,
        "expected_root_document_id": text(resource.get("rootDocumentId")),
        "expected_file_document_id": text(resource.get("fileDocumentId") or file_id),
        "expected_provider_item_id": text(resource.get("providerItemId")),
        "expected_block_id": text(block.get("blockId")),
        "expected_source_path": path,
        "expected_block_order": order,
        "expected_library": "feishu",
        "requested_library": "feishu",
        "expected_role": text(block.get("blockRole")),
        "expected_scope": text(resource.get("permissionScope")),
        "split_fingerprint": fingerprint,
        "split_group": f"feishu:{file_id}:{fingerprint[:16]}",
        "document_title": text(resource.get("title")),
        "block_order": order,
        "source_excerpt": normalized(block.get("normalizedText") or block.get("rawText"))[:600],
        "resolved_section": text(block.get("section")),
        "resolved_block_type": text(block.get("blockType")),
        "manual_score": int(case.get("manual_score", 8)),
        "score_reason": case.get("score_reason", "目标章节直接覆盖 query 的核心知识点，且保留了必要条件或公式上下文。"),
        "retrieval_status": "pending_real_http_run",
        "retrieved_hits": [],
        "exact_block_hit": None,
        "physical_file_hit": None,
        "same_file_hit": None,
        "evidence_window_hit": None,
    })
    return expected


def yaml_scalar(value):
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return json.dumps(str(value), ensure_ascii=False)


def write_yaml(payload, path):
    lines = [f"datasetVersion: {yaml_scalar(payload['datasetVersion'])}",
             f"schemaVersion: {yaml_scalar(payload['schemaVersion'])}",
             f"description: {yaml_scalar(payload['description'])}", "cases:"]
    for case in payload["cases"]:
        lines.append("  -")
        for key, value in case.items():
            if isinstance(value, list):
                lines.append(f"    {key}:")
                for item in value:
                    lines.append(f"      - {yaml_scalar(item)}")
            else:
                lines.append(f"    {key}: {yaml_scalar(value)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(payload, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    headers = ["case_id", "topic", "query_intent", "query", "expected_source_path", "rootDocumentId",
               "fileDocumentId", "providerItemId", "blockId", "blockOrder", "splitFingerprint", "targetSection",
               "resolvedSection", "sourceExcerpt", "manualSummary", "acceptableEvidence", "manualScore",
               "scoreReason", "retrievalStatus", "physicalFileHit", "exactBlockHit", "sameFileHit", "evidenceWindowHit"]
    ws.append(["人工高中数学 FILE 检索标注集", None])
    ws.append(["schemaVersion", payload["schemaVersion"]])
    ws.append([])
    ws.append(headers)
    for case in payload["cases"]:
        ws.append([case.get("case_id"), case.get("topic"), case.get("query_intent"), case.get("query"),
                   case.get("expected_source_path"), case.get("expected_root_document_id"),
                   case.get("expected_file_document_id"), case.get("expected_provider_item_id"),
                   case.get("expected_block_id"), case.get("expected_block_order"), case.get("split_fingerprint"),
                   case.get("target_section"), case.get("resolved_section"), case.get("source_excerpt"),
                   case.get("manual_summary"), "；".join(case.get("acceptable_evidence", [])), case.get("manual_score"),
                   case.get("score_reason"), case.get("retrieval_status"), case.get("physical_file_hit"),
                   case.get("exact_block_hit"), case.get("same_file_hit"), case.get("evidence_window_hit")])
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:W{ws.max_row}"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True, size=13)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 22, 18, 42, 48, 22, 22, 22, 22, 11, 68, 28, 28, 75, 56, 38, 11, 42, 22, 14, 14, 14, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 72
    ws.sheet_properties.tabColor = "1F4E78"
    ws2 = wb.create_sheet("README")
    ws2.append(["说明", payload["description"]])
    ws2.append(["人工来源", "query、topic、target_section、manual_summary、acceptable_evidence 由人工阅读真实正文后填写。"])
    ws2.append(["绑定规则", "source_path + block_order 仅用于绑定当前真实快照；rootDocumentId 与 fileDocumentId 分开保存。"])
    ws2.append(["检索结果", "尚未运行真实 HTTP/GPU 评测前保持 pending，不把预期字段伪装成实际结果。"])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 120
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main():
    payload = json.loads(ANNOTATED.read_text(encoding="utf-8"))
    snapshot = json.loads(SNAPSHOT.read_text(encoding="utf-8"))
    by_path = {text(item.get("sourcePath")): item for item in snapshot["resources"]}
    blocks_by_doc = {text(doc_id): blocks for doc_id, blocks in snapshot["blocks"].items()}
    existing = {case["case_id"]: case for case in payload["cases"]}
    # Correct the one inventory path found during real snapshot validation.
    for case in existing.values():
        if case["case_id"] == "CP-004":
            case["expected_source_path"] = "排列组合知识点/基本的选座问题.md"
            case["expected_block_order"] = 0
    for case_id, correction in REVIEWED_CORRECTIONS.items():
        if case_id not in existing:
            raise RuntimeError(f"missing reviewed case for correction: {case_id}")
        # Keep the query human-authored while binding it to a textual block.
        existing[case_id].update(correction)
    for case in ADDITIONS:
        # Re-apply the reviewed manual definition so the exporter is idempotent.
        existing[case["case_id"]] = case
    cases = [bind_case(existing[key], by_path, blocks_by_doc, index) for index, key in enumerate(sorted(existing), start=1)]
    payload["datasetVersion"] = "teacher-math-manual-annotated-20260827-v2"
    payload["description"] = "人工阅读真实 Feishu 高中数学 Markdown 后编写的 FILE-scoped 检索集；当前版本含人工 query、知识点、目标章节、来源摘录和真实 FILE/block 身份绑定。"
    payload["caseCount"] = len(cases)
    payload["positiveOnly"] = True
    payload["negativeCases"] = 0
    payload["cases"] = cases
    payload["binding"] = {"snapshot": str(SNAPSHOT.relative_to(ROOT)), "status": "bound_to_current_snapshot", "sourcePathAndBlockOrderOnlyForBinding": True}
    ANNOTATED.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_yaml(payload, OUT / "teacher_math_manual_annotated_20260827.yaml")
    write_xlsx(payload, OUT / "teacher_math_manual_annotated_20260827.xlsx")
    print(json.dumps({"caseCount": len(cases), "uniqueFiles": len({x['expected_file_document_id'] for x in cases}), "topics": sorted({x['topic'].split('/')[0] for x in cases})}, ensure_ascii=False))


if __name__ == "__main__":
    main()
